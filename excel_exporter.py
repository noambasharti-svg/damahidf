import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime
import io

def generate_damah_excel(reports, date_str):
    """
    Generates an Excel workbook in memory matching 'דמח אכא יוני 26 פורמט תעצ.xlsx'
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Format sheet name based on date (e.g. דמח 22.7.26)
    try:
        dt = datetime.datetime.strptime(date_str, '%Y-%m-%d')
        sheet_title = f"דמח {dt.day}.{dt.month}.{str(dt.year)[-2:]}"
    except Exception:
        sheet_title = "דמח"
        
    ws.title = sheet_title
    ws.sheet_view.rightToLeft = True
    
    # Styling definitions
    header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    
    data_font = Font(name="Arial", size=11)
    total_font = Font(name="Arial", size=12, bold=True)
    total_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    total_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='double', color='000000')
    )

    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_right = Alignment(horizontal='right', vertical='center')

    # Headers
    headers = [
        'מס״ד',
        'סמכות',
        'יחידה מסגרת',
        'מצבות כ"א סה"כ',
        'התייצבו',
        'מילואים',
        'עבודה מהבית',
        'רידוד סד"כ',
        'לא נוכח מסיבות אחרות(מחלה/חופשה וכו)'
    ]
    
    ws.row_dimensions[1].height = 32
    for col_num, header_text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header_text
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center

    # Data Rows
    current_row = 2
    for report in reports:
        ws.row_dimensions[current_row].height = 22
        
        sid = report.get('sid') or (current_row - 1)
        authority = report.get('authority', 'אכ"א')
        unit_name = report.get('unit_name', '')
        quota = report.get('quota', 0)
        
        is_sub = report.get('is_submitted', 0)
        present = report.get('present_base') if is_sub else None
        reserve = report.get('reserve') if is_sub else None
        wfh = report.get('work_from_home') if is_sub else None
        standby = report.get('standby_reduction') if is_sub else None
        other = report.get('other_absent') if is_sub else None
        
        row_vals = [
            sid,
            authority,
            unit_name,
            quota,
            present,
            reserve,
            wfh,
            standby,
            other
        ]
        
        for col_num, val in enumerate(row_vals, 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.value = val
            cell.font = data_font
            cell.border = thin_border
            if col_num in [1, 2, 4, 5, 6, 7, 8, 9]:
                cell.alignment = align_center
            else:
                cell.alignment = align_right
                
        # Column J formula for row total check
        sum_cell = ws.cell(row=current_row, column=10)
        sum_cell.value = f"=E{current_row}+F{current_row}+G{current_row}+H{current_row}+I{current_row}"
        sum_cell.font = data_font
        sum_cell.alignment = align_center
        sum_cell.border = thin_border
        
        current_row += 1
        
    # Total Summary Row
    ws.row_dimensions[current_row].height = 26
    ws.cell(row=current_row, column=3, value='סה"כ').alignment = align_right
    
    # Formulas for columns D through J
    cols = ['D', 'E', 'F', 'G', 'H', 'I', 'J']
    for idx, col in enumerate(cols, 4):
        cell = ws.cell(row=current_row, column=idx)
        cell.value = f"=SUM({col}2:{col}{current_row - 1})"
        cell.alignment = align_center
        
    for c in range(1, 11):
        cell = ws.cell(row=current_row, column=c)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = total_border

    # Adjust Column Widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if cell.row == 1:
                # header text length
                max_len = max(max_len, len(val_str))
            else:
                max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    ws.column_dimensions['C'].width = 24  # Unit Name
    ws.column_dimensions['I'].width = 30  # Other absent text header
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
